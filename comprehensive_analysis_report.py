#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HVDC Warehouse & Site Material Analysis Script
- Excel 보고서 중심의 종합적인 물류 분석 시스템
- KPI, 누적재고, 도달률, Dead Stock, 회전율 자동 분석
- 시각화 및 엑셀 리포트 자동 생성
"""

import pandas as pd
import numpy as np
from collections import Counter
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
import warnings
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import os
import sys

warnings.filterwarnings('ignore')

# === Configuration ===
excel_path = 'data/HVDC WAREHOUSE_HITACHI(HE).xlsx'  # 분석 대상 데이터 엑셀 파일 경로
sheet_name = 'CASE LIST'                       # 데이터 시트명
start_date = '2023-01-01'                      # 분석 시작 기간
end_date = '2025-12-31'                        # 분석 종료 기간
EMBED_IMAGES = True                            # 생성된 그래프를 엑셀 리포트에 첨부할지 여부

def main():
    """메인 분석 함수"""
    print("=== HVDC Warehouse & Site Material Analysis System ===")
    print(f"실행 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 1. 데이터 로드
    print(f"\n📁 데이터 파일 로드: {excel_path}")
    if not os.path.exists(excel_path):
        print(f"❌ 오류: 데이터 파일을 찾을 수 없습니다: {excel_path}")
        return
    
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    print(f"✅ 데이터 로드 완료 (총 {len(df)}행)")

    # 2. 데이터 전처리
    print("\n🔧 데이터 전처리 중...")
    df['Case No.'] = df['Case No.'].astype(str)        # Case No.을 문자열로 변환
    
    # 날짜 컬럼들을 datetime으로 변환 (Case No.와 Site 컬럼 제외)
    date_cols = [col for col in df.columns if col not in ['Case No.', 'Site']]
    for col in date_cols:
        try:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        except Exception:
            pass

    # 3. 창고 및 현장 컬럼 식별
    warehouse_cols = ['DSV Outdoor', 'DSV Indoor', 'DSV Al Markaz', 'Hauler Indoor', 'DSV MZP', 'MOSB']
    site_cols = ['DAS', 'MIR', 'SHU', 'AGI']
    
    # 혹시 대비: df에 존재하지 않는 컬럼은 제거
    warehouse_cols = [col for col in warehouse_cols if col in df.columns]
    site_cols = [col for col in site_cols if col in df.columns]

    # 'Site' 컬럼에 "Das"와 같이 소문자 표기된 값이 있다면 대문자로 통일
    if 'Site' in df.columns:
        df['Site'] = df['Site'].replace({'Das': 'DAS', 'Mir': 'MIR', 'Shu': 'SHU', 'Agi': 'AGI'})

    print(f"🏭 입고(창고) 컬럼: {warehouse_cols}")
    print(f"🏗️  출고(현장) 컬럼: {site_cols}")

    # 4. 월별 범위 생성 (월말 기준)
    months = pd.date_range(start=start_date, end=end_date, freq='M')
    month_labels = [m.strftime("%Y-%m") for m in months]

    # 5. 개선된 월별 입출고 이벤트 계산
    print("\n📊 월별 입출고 이벤트 집계 중...")
    warehouse_stats = {wh: {'in': Counter(), 'out': Counter()} for wh in warehouse_cols}
    site_stats = {site: Counter() for site in site_cols}

    for _, row in df.iterrows():
        events = []
        # 창고 입고 이벤트 수집
        for wh in warehouse_cols:
            if pd.notna(row[wh]):
                events.append((pd.to_datetime(row[wh]), wh, 'warehouse_in'))
        # 현장 수령(출고) 이벤트 수집
        for site in site_cols:
            if pd.notna(row[site]):
                events.append((pd.to_datetime(row[site]), site, 'site_out'))
        if not events:
            continue
        events.sort(key=lambda x: x[0])  # 시간 순 정렬
        
        prev_loc = None
        for date, loc, ev_type in events:
            month_str = date.strftime("%Y-%m")
            if ev_type == 'warehouse_in':
                if prev_loc is None:
                    # 외부 -> 창고 첫 입고
                    warehouse_stats[loc]['in'][month_str] += 1
                else:
                    # 창고 간 이동: 이전 창고에서 출고, 새로운 창고에 입고
                    warehouse_stats[prev_loc]['out'][month_str] += 1
                    warehouse_stats[loc]['in'][month_str] += 1
                prev_loc = loc  # 현재 위치를 prev_loc로 설정
            elif ev_type == 'site_out':
                # 현장 최종 출고 이벤트
                if prev_loc is not None:
                    warehouse_stats[prev_loc]['out'][month_str] += 1
                site_stats[loc][month_str] += 1
                prev_loc = None  # 출고 후 재고로 남지 않음

    print("✅ 월별 입출고 이벤트 집계 완료")

    # 6. 월별 재고 계산 (창고별 재고 및 현장별 누적 재고)
    print("\n📈 월별 재고 계산 중...")
    warehouse_stock = {}
    site_stock = {}

    for wh in warehouse_cols:
        stock_data = []
        current_stock = 0
        for month in month_labels:
            inbound = warehouse_stats[wh]['in'].get(month, 0)
            outbound = warehouse_stats[wh]['out'].get(month, 0)
            current_stock += inbound - outbound
            stock_data.append({'월': month, '입고': inbound, '출고': outbound, '재고': current_stock})
        warehouse_stock[wh] = pd.DataFrame(stock_data)

    for site in site_cols:
        stock_data = []
        cumulative = 0
        for month in month_labels:
            inbound = site_stats[site].get(month, 0)
            cumulative += inbound
            stock_data.append({'월': month, '입고': inbound, '누적재고': cumulative})
        site_stock[site] = pd.DataFrame(stock_data)

    print("✅ 월별 재고 계산 완료")

    # 7. KPI 계산: Site별 도달률(%) 및 평균 리드타임
    print("\n📋 Site별 KPI 계산 중...")
    total_cases = df['Case No.'].nunique()
    site_kpi_list = []
    for site in site_cols:
        temp = df[df[site].notna()].copy()
        reached = temp['Case No.'].nunique()
        reach_rate = round((reached / total_cases) * 100, 2) if total_cases > 0 else 0.0
        # 각 케이스의 최초 입고일 및 해당 Site 도착 리드타임 계산
        temp['입고일'] = temp[warehouse_cols].min(axis=1)
        temp['리드타임(일)'] = (temp[site] - temp['입고일']).dt.days
        avg_leadtime = round(temp['리드타임(일)'].mean(), 1) if len(temp) > 0 else 0.0
        site_kpi_list.append({'Site': site, '도달건수': reached, '도달률(%)': reach_rate, '평균 리드타임(일)': avg_leadtime})
    site_kpi_df = pd.DataFrame(site_kpi_list)
    print("✅ Site별 KPI 계산 완료")

    # 8. Dead Stock (90일 이상 미출고) 분석
    print("\n🚨 Dead Stock 분석 중...")
    today = pd.Timestamp(datetime.today().strftime('%Y-%m-%d'))
    dead_stock_list = []
    for _, row in df.iterrows():
        # 출고(현장 인도)된 적 없는 케이스 필터
        if not any(pd.notna(row[site]) for site in site_cols):
            # 해당 케이스의 마지막 입고일과 위치 찾기
            inbound_dates = {wh: row[wh] for wh in warehouse_cols if pd.notna(row[wh])}
            if not inbound_dates:
                continue  # 입고 이벤트 자체가 없으면 제외
            # 마지막 입고일 및 위치
            last_loc, last_date = None, pd.NaT
            for wh, date in inbound_dates.items():
                if pd.isna(last_date) or date > last_date:
                    last_date = date
                    last_loc = wh
            days_since = (today - pd.to_datetime(last_date)).days
            if days_since > 90:
                dead_stock_list.append({
                    'Case No.': row['Case No.'],
                    '마지막입고일': last_date,
                    '마지막위치': last_loc,
                    '입고후경과일': days_since,
                    'Site': row.get('Site', '')
                })
    dead_stock_df = pd.DataFrame(dead_stock_list)
    print(f"✅ Dead Stock (입고 후 90일 이상 미출고) 건수: {len(dead_stock_df)}건")

    # 9. 창고별 월별 회전율 분석 (월별 출고/입고 비율)
    print("\n🔄 창고별 회전율 계산 중...")
    turnover_dict = {}
    for wh in warehouse_cols:
        df_wh = df[df[wh].notna()].copy()
        df_wh['입고월'] = pd.to_datetime(df_wh[wh]).dt.to_period('M').dt.to_timestamp() + pd.offsets.MonthEnd(0)
        # 해당 창고로 입고된 케이스 중 최종 출고된 비율 계산
        df_wh['출고여부'] = df_wh[site_cols].notna().any(axis=1)
        monthly_in = df_wh.groupby('입고월')['Case No.'].count()
        monthly_out = df_wh[df_wh['출고여부']].groupby('입고월')['Case No.'].count()
        turnover = (monthly_out / monthly_in).fillna(0).clip(upper=1)
        # 월 범위에 맞춰 인덱스 보정 (분석 기간 내 월별 데이터 보완)
        turnover_full = turnover.reindex(months, fill_value=0)
        turnover_dict[wh] = turnover_full

    # 회전율 데이터프레임 생성 (index: 월, columns: 각 창고 회전율)
    turnover_df = pd.DataFrame(turnover_dict)
    # 인덱스를 "YYYY-MM" 문자열로 변환하여 보기 쉽게 설정
    turnover_df.index = turnover_df.index.strftime("%Y-%m")
    turnover_df.index.name = '월'
    turnover_df = turnover_df.round(3)  # 소수 셋째자리 정도까지 표시
    print("✅ 창고별 회전율 계산 완료")

    # 10. 시각화: Site별 리드타임 분포 히스토그램
    print("\n📊 시각화 생성 중...")
    leadtime_all = []
    for site in site_cols:
        temp = df[df[site].notna()].copy()
        temp['입고일'] = temp[warehouse_cols].min(axis=1)
        temp['리드타임(일)'] = (temp[site] - temp['입고일']).dt.days
        temp = temp[['리드타임(일)']].dropna()
        temp['Site'] = site
        leadtime_all.append(temp)
    leadtime_df = pd.concat(leadtime_all, ignore_index=True)
    
    plt.figure(figsize=(12, 8))
    sns.histplot(data=leadtime_df, x='리드타임(일)', hue='Site', bins=30, multiple='stack', kde=True)
    plt.title('Site별 리드타임 분포 (일)', fontsize=16, fontweight='bold')
    plt.xlabel('리드타임 (일)', fontsize=12)
    plt.ylabel('케이스 수', fontsize=12)
    plt.grid(True, alpha=0.3)
    plt.legend(title='Site', title_fontsize=12)
    plt.tight_layout()
    
    # outputs 폴더 생성
    output_dir = 'outputs'
    os.makedirs(output_dir, exist_ok=True)
    hist_image_path = os.path.join(output_dir, 'site_leadtime_distribution.png')
    plt.savefig(hist_image_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"✅ 리드타임 분포 히스토그램 생성 및 저장: {hist_image_path}")

    # 11. 시각화: 창고별 월별 회전율 추이 라인 차트
    plt.figure(figsize=(14, 8))
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b']
    for i, wh in enumerate(warehouse_cols):
        plt.plot(month_labels, turnover_df[wh].values, marker='o', label=wh, 
                linewidth=2, markersize=6, color=colors[i % len(colors)])
    plt.title('창고별 월별 회전율 추이', fontsize=16, fontweight='bold')
    plt.xlabel('월', fontsize=12)
    plt.ylabel('회전율 (출고/입고)', fontsize=12)
    plt.xticks(rotation=45)
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    
    line_image_path = os.path.join(output_dir, 'warehouse_turnover.png')
    plt.savefig(line_image_path, dpi=300, bbox_inches='tight')
    plt.close()
    print(f"✅ 창고별 회전율 라인 차트 생성 및 저장: {line_image_path}")

    # 12. 엑셀로 결과 저장
    print(f"\n📋 엑셀 리포트 생성 중...")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'HVDC_Warehouse_Analysis_Report_{timestamp}.xlsx')
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 창고별 월별 입출고/재고 (각 창고별 시트)
        for wh, df_stock in warehouse_stock.items():
            # 총합 행 추가 (입고/출고 합계, 재고는 마지막 재고 합산)
            sums = df_stock[['입고', '출고']].sum()  # '재고'는 합산하지 않음
            sums['재고'] = df_stock['재고'].iloc[-1]  # 현재 재고 (마지막 값)
            total_row = pd.DataFrame([sums], index=['총합'])
            df_out = pd.concat([df_stock, total_row])
            df_out.to_excel(writer, sheet_name=f'창고_{wh}')
        
        # 현장별 월별 입고/누적재고 (각 Site별 시트)
        for site, df_stock in site_stock.items():
            sums = df_stock[['입고']].sum()
            sums['누적재고'] = df_stock['누적재고'].iloc[-1]  # 현재 누적재고
            total_row = pd.DataFrame([sums], index=['총합'])
            df_out = pd.concat([df_stock, total_row])
            df_out.to_excel(writer, sheet_name=f'Site_{site}')
        
        # Dead Stock 목록 시트
        if len(dead_stock_df) > 0:
            # 날짜 형식 변환 (yyyy-mm-dd)
            dead_stock_df['마지막입고일'] = dead_stock_df['마지막입고일'].dt.strftime('%Y-%m-%d')
        dead_stock_df.to_excel(writer, sheet_name='DeadStock_90일+', index=False)
        
        # Site KPI 시트
        site_kpi_df.to_excel(writer, sheet_name='Site_KPI', index=False)
        
        # 창고별 회전율 시트
        turnover_df.to_excel(writer, sheet_name='회전율')  # '월' 인덱스 포함 출력

    # 13. 엑셀에 그래프 이미지 삽입 (옵션)
    if EMBED_IMAGES:
        print("🖼️  그래프 이미지 엑셀 삽입 중...")
        wb = load_workbook(output_path)
        # 시각화 전용 시트 생성
        ws = wb.create_sheet('시각화')
        try:
            img1 = ExcelImage(hist_image_path)
            img2 = ExcelImage(line_image_path)
            # 이미지 크기를 적절히 조절
            img1.width, img1.height = img1.width * 0.6, img1.height * 0.6
            img2.width, img2.height = img2.width * 0.6, img2.height * 0.6
            ws.add_image(img1, 'A1')
            ws.add_image(img2, 'A25')
        except Exception as e:
            print(f"⚠️ 이미지 삽입 실패: {e}")
        wb.save(output_path)
        wb.close()

    print(f"\n✅ 분석 완료! 엑셀 리포트 저장: {output_path}")
    print(f"📄 파일 크기: {os.path.getsize(output_path) / 1024:.1f} KB")
    
    print("\n📋 생성된 엑셀 파일에 다음 시트들이 포함되어 있습니다:")
    print("- 창고별 월별 입출고/재고 (창고별 개별 시트)")
    print("- 현장별 월별 입고/누적재고 (Site별 개별 시트)")
    print("- Site_KPI (현장별 도달률 및 평균 리드타임)")
    print("- DeadStock_90일+ (90일 넘게 미출고된 Dead Stock 목록)")
    print("- 회전율 (창고별 월별 회전율 데이터)")
    if EMBED_IMAGES:
        print("- 시각화 (분석 차트 이미지)")
    
    # 14. 엑셀 파일 자동 열기
    try:
        os.startfile(output_path)
        print(f"\n🔓 엑셀 파일이 자동으로 열렸습니다.")
    except:
        print(f"\n💡 엑셀 파일을 수동으로 열어주세요: {output_path}")

if __name__ == "__main__":
    main() 